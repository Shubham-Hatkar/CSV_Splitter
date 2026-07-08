

import csv
import io
import json
import re
import shutil
import time
import zipfile
from datetime import date, datetime
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
STORAGE_DIR = APP_DIR / "storage"
UPLOADS_DIR = STORAGE_DIR / "uploads"
OUTPUTS_DIR = STORAGE_DIR / "outputs"
LOGS_DIR = STORAGE_DIR / "logs"
TEMP_DIR = STORAGE_DIR / "temp"
HISTORY_FILE = STORAGE_DIR / "job_history.json"
DEFAULT_HEADER_ROWS = 5
DEFAULT_PROGRESS_EVERY = 50000


def ensure_storage():
    for folder in [UPLOADS_DIR, OUTPUTS_DIR, LOGS_DIR, TEMP_DIR]:
        folder.mkdir(parents=True, exist_ok=True)
    if not HISTORY_FILE.exists():
        HISTORY_FILE.write_text("[]", encoding="utf-8")


def parse_size(size_text):
    text = str(size_text).strip().upper().replace(" ", "")
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(B|KB|MB|GB)?", text)
    if not match:
        raise ValueError("Use values like 500KB, 10MB, 90MB, or 1048576")
    number = float(match.group(1))
    unit = match.group(2) or "B"
    multipliers = {"B": 1, "KB": 1024, "MB": 1024**2, "GB": 1024**3}
    size_bytes = int(number * multipliers[unit])
    if size_bytes <= 0:
        raise ValueError("Size must be greater than zero")
    return size_bytes


def sanitize_name(value):
    text = str(value).strip() or "file"
    text = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text[:120]


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def detect_csv_dialect(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(65536)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class SimpleDialect(csv.Dialect):
            delimiter = ","
            quotechar = '"'
            doublequote = True
            skipinitialspace = False
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL
        return SimpleDialect


def reconcile_row(row, expected_width, delimiter):
    if len(row) == expected_width:
        return row, False
    if len(row) < expected_width:
        return row + [""] * (expected_width - len(row)), True
    merged_last = row[: expected_width - 1] + [delimiter.join(row[expected_width - 1:])]
    return merged_last, True


def get_sheet_names(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def preview_headers(file_bytes, ext, keep_header_rows, sheet_name=None):
    rows = []
    if ext == ".csv":
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text), detect_csv_dialect_from_text(text))
        for _ in range(keep_header_rows):
            rows.append([normalize_cell(x) for x in next(reader)])
        return rows

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    row_iter = ws.iter_rows(values_only=True)
    for _ in range(keep_header_rows):
        rows.append([normalize_cell(x) for x in next(row_iter)])
    wb.close()
    return rows


def detect_csv_dialect_from_text(text):
    sample = text[:65536]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class SimpleDialect(csv.Dialect):
            delimiter = ","
            quotechar = '"'
            doublequote = True
            skipinitialspace = False
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL
        return SimpleDialect


def csv_source(path, keep_header_rows):
    f = open(path, "r", encoding="utf-8-sig", newline="")
    dialect = detect_csv_dialect(path)
    reader = csv.reader(f, dialect)
    header_rows = [[normalize_cell(x) for x in next(reader)] for _ in range(keep_header_rows)]
    expected_width = len(header_rows[-1])
    stats = {"adjusted_rows": 0}

    def iterator():
        for row_num, row in enumerate(reader, start=keep_header_rows + 1):
            fixed, changed = reconcile_row([normalize_cell(x) for x in row], expected_width, dialect.delimiter)
            if changed:
                stats["adjusted_rows"] += 1
            yield row_num, fixed

    return header_rows, iterator(), f.close, stats


def excel_source(path, keep_header_rows, sheet_name=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(values_only=True)
    header_rows = [[normalize_cell(x) for x in next(rows)] for _ in range(keep_header_rows)]
    expected_width = len(header_rows[-1])
    stats = {"adjusted_rows": 0}

    def iterator():
        for row_num, row in enumerate(rows, start=keep_header_rows + 1):
            fixed, changed = reconcile_row([normalize_cell(x) for x in row], expected_width, ",")
            if changed:
                stats["adjusted_rows"] += 1
            yield row_num, fixed

    return header_rows, iterator(), wb.close, stats


class CsvSizer:
    def __init__(self):
        self.buffer = io.StringIO()
        self.writer = csv.writer(self.buffer, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")

    def row_size(self, row):
        self.buffer.seek(0)
        self.buffer.truncate(0)
        self.writer.writerow(row)
        return len(self.buffer.getvalue().encode("utf-8"))


class DiskCsvSplitter:
    def __init__(self, temp_dir, base_name, header_rows, rows_per_file=None, max_size_bytes=None):
        self.temp_dir = Path(temp_dir)
        self.base_name = sanitize_name(base_name)
        self.header_rows = header_rows
        self.rows_per_file = rows_per_file
        self.max_size_bytes = max_size_bytes
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.sizer = CsvSizer()
        self.header_size = sum(self.sizer.row_size(r) for r in header_rows)
        self.part_number = 0
        self.files_created = 0
        self.total_rows_written = 0
        self.current_file = None
        self.current_writer = None
        self.current_size = 0
        self.current_rows = 0

    def open_new_file(self):
        self.part_number += 1
        self.files_created += 1
        path = self.temp_dir / f"{self.base_name}_part_{self.part_number:03d}.csv"
        self.current_file = open(path, "w", encoding="utf-8", newline="")
        self.current_writer = csv.writer(self.current_file, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        for row in self.header_rows:
            self.current_writer.writerow(row)
        self.current_size = self.header_size
        self.current_rows = 0

    def close_current(self):
        if self.current_file:
            self.current_file.close()
            self.current_file = None
            self.current_writer = None

    def write_row(self, row):
        if self.current_file is None:
            self.open_new_file()
        row_size = self.sizer.row_size(row)
        if self.rows_per_file and self.current_rows >= self.rows_per_file:
            self.close_current()
            self.open_new_file()
        if self.max_size_bytes and self.current_rows > 0 and self.current_size + row_size > self.max_size_bytes:
            self.close_current()
            self.open_new_file()
        self.current_writer.writerow(row)
        self.current_rows += 1
        self.current_size += row_size
        self.total_rows_written += 1

    def close(self):
        self.close_current()


def write_log(log_path, message):
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(message + "\n")


def save_history(entry):
    history = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    history.insert(0, entry)
    HISTORY_FILE.write_text(json.dumps(history[:100], indent=2), encoding="utf-8")


def load_history():
    return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))


def zip_folder(folder_path, zip_path):
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_path in sorted(Path(folder_path).glob("*.csv")):
            zf.write(file_path, arcname=file_path.name)


def process_file(input_path, ext, split_mode, rows_per_file, max_size_bytes, keep_header_rows,
                 sheet_name, filter_column, filter_value, progress_every, status_box):
    job_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = sanitize_name(input_path.stem)
    temp_job_dir = TEMP_DIR / job_id
    log_path = LOGS_DIR / f"{job_id}_{base_name}.log"
    zip_path = OUTPUTS_DIR / f"{job_id}_{base_name}_split.zip"

    source_fn = csv_source if ext == ".csv" else excel_source
    header_rows, rows, cleanup, stats = source_fn(input_path, keep_header_rows) if ext == ".csv" else source_fn(input_path, keep_header_rows, sheet_name)

    processed_rows = 0
    written_rows = 0
    filter_index = None
    started = time.perf_counter()
    write_log(log_path, f"Started job {job_id} for {input_path.name}")

    try:
        if filter_column:
            filter_index = header_rows[-1].index(filter_column)

        splitter = DiskCsvSplitter(
            temp_dir=temp_job_dir,
            base_name=base_name,
            header_rows=header_rows,
            rows_per_file=rows_per_file if split_mode == "rows" else None,
            max_size_bytes=max_size_bytes if split_mode == "size" else None,
        )

        for _, row in rows:
            processed_rows += 1
            if filter_index is not None and row[filter_index] != filter_value:
                pass
            else:
                splitter.write_row(row)
                written_rows += 1

            if processed_rows % progress_every == 0:
                elapsed = max(time.perf_counter() - started, 0.001)
                msg = (
                    f"Processed={processed_rows:,} | Written={written_rows:,} | "
                    f"Files={splitter.files_created:,} | Elapsed={elapsed:.1f}s | "
                    f"Rate={processed_rows / elapsed:,.0f} rows/s"
                )
                write_log(log_path, msg)
                status_box.text(msg)

        splitter.close()
        zip_folder(temp_job_dir, zip_path)
        elapsed = max(time.perf_counter() - started, 0.001)
        final_msg = (
            f"Completed | Processed={processed_rows:,} | Written={written_rows:,} | "
            f"Files={splitter.files_created:,} | Adjusted={stats['adjusted_rows']:,} | "
            f"Elapsed={elapsed:.2f}s | Rate={processed_rows / elapsed:,.0f} rows/s"
        )
        write_log(log_path, final_msg)
        status_box.text(final_msg)

        result = {
            "job_id": job_id,
            "file_name": input_path.name,
            "split_mode": split_mode,
            "rows_per_file": rows_per_file if split_mode == "rows" else None,
            "max_size": max_size_bytes if split_mode == "size" else None,
            "processed_rows": processed_rows,
            "written_rows": written_rows,
            "files_created": splitter.files_created,
            "adjusted_rows": stats["adjusted_rows"],
            "elapsed_seconds": round(elapsed, 2),
            "zip_path": str(zip_path),
            "log_path": str(log_path),
            "upload_path": str(input_path),
            "status": "Success",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        save_history(result)
        return result
    except Exception as exc:
        write_log(log_path, f"Failed: {exc}")
        error_result = {
            "job_id": job_id, "file_name": input_path.name, "status": f"Failed: {exc}",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        save_history(error_result)
        raise
    finally:
        cleanup()
        if temp_job_dir.exists():
            shutil.rmtree(temp_job_dir, ignore_errors=True)


ensure_storage()
st.set_page_config(page_title="CSV/XLSX Splitter", layout="wide")
st.title("CSV/XLSX Splitter")
st.caption("Disk-based app with uploads, outputs, logs, and job history.")

uploaded_file = st.file_uploader("Select input file", type=["csv", "xlsx", "xlsm"])
left, right = st.columns(2)

with left:
    split_mode = st.radio("Split by", ["rows", "size"], horizontal=True)
    keep_header_rows = st.number_input("Header rows to keep", min_value=1, value=DEFAULT_HEADER_ROWS, step=1)
    progress_every = st.number_input("Log every N rows", min_value=1000, value=DEFAULT_PROGRESS_EVERY, step=1000)

with right:
    rows_per_file = st.number_input("Rows per file", min_value=1, value=50000, step=1000) if split_mode == "rows" else None
    max_size_text = st.text_input("Max file size", value="90MB") if split_mode == "size" else None
    max_size_bytes = parse_size(max_size_text) if max_size_text else None

sheet_name = None
filter_column = None
filter_value = None
file_bytes = uploaded_file.getvalue() if uploaded_file else None
ext = f".{uploaded_file.name.split('.')[-1].lower()}" if uploaded_file else None

if uploaded_file and ext in [".xlsx", ".xlsm"]:
    sheet_name = st.selectbox("Select sheet", get_sheet_names(file_bytes))

if uploaded_file:
    try:
        header_preview = preview_headers(file_bytes, ext, keep_header_rows, sheet_name)
        with st.expander("Preview header rows", expanded=False):
            for idx, row in enumerate(header_preview, start=1):
                st.write(f"Row {idx}: {row}")

        with st.expander("Optional filter", expanded=False):
            if st.checkbox("Enable filter"):
                filter_column = st.selectbox("Filter column", header_preview[-1])
                filter_value = st.text_input("Filter value")
    except Exception as exc:
        st.error(f"Header preview failed: {exc}")

if st.button("Split file", type="primary", disabled=uploaded_file is None):
    try:
        saved_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{sanitize_name(uploaded_file.name)}"
        upload_path = UPLOADS_DIR / saved_name
        with open(upload_path, "wb") as f:
            f.write(file_bytes)

        status_box = st.empty()
        result = process_file(
            input_path=upload_path,
            ext=ext,
            split_mode=split_mode,
            rows_per_file=rows_per_file,
            max_size_bytes=max_size_bytes,
            keep_header_rows=keep_header_rows,
            sheet_name=sheet_name,
            filter_column=filter_column,
            filter_value=filter_value,
            progress_every=progress_every,
            status_box=status_box,
        )

        st.success("File processed successfully.")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Processed", f"{result['processed_rows']:,}")
        c2.metric("Written", f"{result['written_rows']:,}")
        c3.metric("Files created", f"{result['files_created']:,}")
        c4.metric("Elapsed", f"{result['elapsed_seconds']:.2f}s")

        st.write("**Stored files**")
        st.write(f"Upload: `{result['upload_path']}`")
        st.write(f"ZIP output: `{result['zip_path']}`")
        st.write(f"Log file: `{result['log_path']}`")

        with open(result["zip_path"], "rb") as f:
            st.download_button("Download ZIP", data=f.read(), file_name=Path(result["zip_path"]).name, mime="application/zip")

        st.text_area("Run log", value=Path(result["log_path"]).read_text(encoding="utf-8"), height=220)
    except Exception as exc:
        st.error(f"Error: {exc}")

st.divider()
st.subheader("Job History")
history = load_history()

if not history:
    st.info("No jobs yet.")
else:
    for item in history[:10]:
        cols = st.columns([2, 2, 2, 2, 3])
        cols[0].write(item.get("created_at", ""))
        cols[1].write(item.get("file_name", ""))
        cols[2].write(item.get("status", ""))
        cols[3].write(str(item.get("processed_rows", "")))
        cols[4].write(item.get("zip_path", ""))
